import os
import re
from datetime import datetime
from flask import Flask, request, render_template, session, send_file, jsonify
from werkzeug.utils import secure_filename
import PyPDF2
import pandas as pd
from openpyxl import load_workbook
import tempfile
import shutil
from config import *

"""
WELL Certification PDF Parser with Robust Scoring Rules

Parsing Rules:
1. Parse using robust regex that handles decimals and sums like 1+1, 0.5+1
2. Format: code: A05.1, title: text, pts_attempted: number/number+number/No, 
   status: Achieved|Not Attempted|Pending Documentation|Not Applicable, 
   pts_achieved: optional number at end

Scoring Logic:
- "Pending Documentation" → write "Pending Documentation" (no score)
- "Not Applicable" → write "Not Applicable" 
- "Not Attempted" → write empty
- "Achieved":
  * If trailing pts_achieved exists → write that number (including 0, 0.5, 1+1→2)
  * Else if pts_attempted is numeric → write that number
  * Else → write "p"
- Important: Keep 0 and decimals as numbers (0.5 stays 0.5, don't coerce to blank)

Excel Writing:
- Numeric values written as numbers (float(value)), not text
- Text values (statuses, 'p') written as strings
- Ensures proper Excel formatting and calculations
"""

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# Ensure directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

# Template path constant
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template1.xlsx")  # put your real template here

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_first_two_pages(pdf_path):
    """Extract first 2 pages from PDF using PyPDF2"""
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            if len(pdf_reader.pages) == 0:
                return None, "PDF has no pages"
            
            # Create a new PDF writer
            pdf_writer = PyPDF2.PdfWriter()
            
            # Add first page
            pdf_writer.add_page(pdf_reader.pages[0])
            
            # Add second page if it exists
            if len(pdf_reader.pages) > 1:
                pdf_writer.add_page(pdf_reader.pages[1])
            
            # Save the extracted pages
            first_two_pages_path = os.path.join(PROCESSED_FOLDER, f"first_two_pages_{os.path.basename(pdf_path)}")
            with open(first_two_pages_path, 'wb') as output_file:
                pdf_writer.write(output_file)
            
            return first_two_pages_path, None
            
    except Exception as e:
        return None, f"Error extracting pages: {str(e)}"

def convert_to_markdown(pdf_path):
    """Convert PDF to markdown using PyPDF2 text extraction"""
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            markdown_content = []
            markdown_content.append(f"# PDF Document: {os.path.basename(pdf_path)}\n")
            
            for page_num, page in enumerate(pdf_reader.pages, 1):
                text = page.extract_text()
                if text and text.strip():
                    markdown_content.append(f"## Page {page_num}\n")
                    markdown_content.append(text)
                    markdown_content.append("\n")
            
            return '\n'.join(markdown_content), None
            
    except Exception as e:
        return None, f"Error converting to markdown: {str(e)}"

def create_combined_excel(results):
    """Create WELL certification Excel by writing into the REAL template (preserves merged headers)."""
    try:
        if not results:
            return None, "No results to process"

        # 1) Load the real template (do NOT rebuild headers/merges)
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        # 2) Build a 3-row header map so we can find exact columns
        header_rows = [1, 2, 3]
        max_col = ws.max_column
        headers = []
        for col in range(1, max_col + 1):
            triple = tuple((ws.cell(r, col).value if ws.cell(r, col).value is not None else "") for r in header_rows)
            headers.append(triple)

        from openpyxl.styles import Alignment
        from openpyxl.utils import column_index_from_string

        # headers is your 3-row header list: [(row1,row2,row3), ...]

        def col_for_concept_subpoints(letter, concept_name):
            """Find 'Sub-Points' column by header (row3 == 'Sub-Points' and (row1 letter OR row2 name))."""
            for i, t in enumerate(headers, start=1):
                r1 = (str(t[0]).strip() if t[0] else "")
                r2 = (str(t[1]).strip() if t[1] else "")
                r3 = (str(t[2]).strip() if t[2] else "")
                if r3 == "Sub-Points" and (r1 == letter or r2 == concept_name):
                    return i
            return None

        def col_for_concept_pct(letter):
            """Find the % column by header (row3 == '%' and row1 == letter)."""
            for i, t in enumerate(headers, start=1):
                r1 = (str(t[0]).strip() if t[0] else "")
                r3 = (str(t[2]).strip() if t[2] else "")
                if r3 == "%" and r1 == letter:
                    return i
            return None

        # Optional fallback if header lookup ever fails (your sheet letters)
        FALLBACK_SP_COLS = {"A":"AX","W":"BQ","N":"CL","L":"CZ","V":"DX","T":"EO","S":"FE","X":"GE","M":"HA","C":"IS","I":"JD"}
        def sp_col_fallback(ws, letter):
            try:
                return column_index_from_string(FALLBACK_SP_COLS[letter])
            except Exception:
                return None

        CENTER = Alignment(horizontal="center", vertical="center")

        concept_letter_to_name = {
            "A":"Air","W":"Water","N":"Nourishment","L":"Light","V":"Movement",
            "T":"Thermal Comfort","S":"Sound","X":"Materials","M":"Mind","C":"Community","I":"Innovation"
        }

        def find_col(predicate):
            for idx, triple in enumerate(headers, start=1):
                if predicate(triple):
                    return idx
            return None

        def col_for_project_field(name):
            nm = name.strip()
            return find_col(lambda t: any(isinstance(x, str) and x.strip() == nm for x in t))

        def cols_for_part_code(code_prefix):
            # 3rd row header starts with the code like 'A05.1'
            cols = []
            for idx, t in enumerate(headers, start=1):
                third = t[2]
                if isinstance(third, str) and third.startswith(code_prefix):
                    cols.append(idx)
            return cols



        # 3) Find next empty data row (stay BELOW the merged header block)
        row_idx = 4
        while any(ws.cell(row_idx, j).value not in (None, "") for j in range(1, max_col + 1)):
            row_idx += 1

        processed = 0

        for result in results:
            if result.get("status") != "success":
                continue

            parsed = parse_well_markdown(result.get("markdown", ""))
            if not parsed:
                continue

            # ---- Write basic project info
            def set_field(field_name, value):
                c = col_for_project_field(field_name)
                if c is not None and value not in (None, ""):
                    ws.cell(row_idx, c, value)

            set_field("Project Name", parsed.get("project_name"))
            set_field("Project ID", parsed.get("project_id"))
            set_field("Date Certified", parsed.get("date_cert"))

            # ---- Write parts with rules & accumulate subpoints
            # Before the parts loop, start subpoints accumulator:
            subpoints = {k: 0.0 for k in concept_letter_to_name}

            # When writing each part value:
            for part in parsed.get("parts", []):
                code  = part.get("code", "")
                value = part.get("value", "")
                part_cols = cols_for_part_code(code)  # your existing function that finds all columns starting with this code

                for c in part_cols:
                    if isinstance(value, (int, float)):
                        cell = ws.cell(row_idx, c, float(value))
                    elif value in (None, ""):
                        # leave cell empty
                        continue
                    else:
                        cell = ws.cell(row_idx, c, value)
                    cell.alignment = CENTER  # (3) center values, including 'p' and text statuses

                # (1) Only numeric Achieved contributes to Sub-Points
                if isinstance(value, (int, float)):
                    subpoints[code[0]] += float(value)

            # ---- Sub-Points per concept
            # (1) Write Sub-Points per concept, centered
            total_points = 0.0
            for letter, sp in subpoints.items():
                cname = concept_letter_to_name[letter]
                c_sp = col_for_concept_subpoints(letter, cname) or sp_col_fallback(ws, letter)
                if c_sp:
                    cell = ws.cell(row_idx, c_sp, round(sp, 3))
                    cell.alignment = CENTER
                total_points += sp

            # Total Points = sum of Sub-Points, centered
            c_total = col_for_project_field("Total Points")
            if c_total:
                cell = ws.cell(row_idx, c_total, round(total_points, 3))
                cell.alignment = CENTER

            # ---- Percentages (A..I): subpoints / total_points * 100
            # (2) Percent columns with % sign, no decimals, centered
            if total_points > 0:
                for letter, sp in subpoints.items():
                    c_pct = col_for_concept_pct(letter)
                    if c_pct:
                        frac = sp / total_points  # 0.0–1.0
                        cell = ws.cell(row_idx, c_pct, frac)
                        cell.number_format = "0%"   # shows % sign with no decimals
                        cell.alignment = CENTER

            processed += 1
            row_idx += 1

        if processed == 0:
            # Still produce an empty copy of the template (helps debugging on the client)
            out_name = f"well_certification_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_path = os.path.join(PROCESSED_FOLDER, out_name)
            wb.save(excel_path)
            return excel_path, None

        # 4) Save to processed folder
        out_name = f"well_certification_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_path = os.path.join(PROCESSED_FOLDER, out_name)
        wb.save(excel_path)
        return excel_path, None

    except Exception as e:
        return None, f"Error creating WELL certification Excel: {str(e)}"

def parse_well_markdown(markdown_content):
    try:
        def norm(s: str) -> str:
            s2 = s.replace("β", "")
            s2 = re.sub(r"(\b[AWNLVTSXMCI]\d{2})\.\s+(\d)\b", r"\1.\2", s2)      # A01. 1 -> A01.1
            s2 = re.sub(r"(\d)\s*\.\s*(\d)", r"\1.\2", s2)                        # 0. 5 -> 0.5
            # Only squeeze spaces/tabs, keep newlines for proper line-by-line parsing
            s2 = re.sub(r'[ \t]+', ' ', s2)
            return s2

        text = norm(markdown_content)
        
        # Handle OCR variations of "Pending Documentation & On-Site"
        text = text.replace("Pending Documentation & On-Site", "Pending Documentation")
        text = text.replace("Pending Documentation & On Site", "Pending Documentation")

        # keep line breaks; only squeeze runs of spaces/tabs
        text = re.sub(r'[ \t]+', ' ', text)

        # Basic fields - Accept 9–12 digits; keep leading zeros
        m = re.search(r"\b(\d{9,12})\b\s*-\s*([^-()]+?)(?=\s*\(WELL|\s*Date:|\s*WELL|\s*$)", text)
        if m:
            project_id  = m.group(1)                # e.g., "02202255386"
            project_name = m.group(2).strip()       # e.g., "SAP Labs China, Shanghai Campus"
        else:
            # Fallbacks if OCR is messy: try looser name grab after a long digit cluster
            mid = re.search(r"\b(\d{9,12})\b\s*-\s*", text)
            project_id = mid.group(1) if mid else "Unknown"
            # Try to slice a reasonable name region
            pname = re.search(r"\b\d{9,12}\b\s*-\s*(.*?)(?:\s*Date:|\s*WELL|\s*\(|$)", text)
            project_name = pname.group(1).strip() if pname else "Unknown Project"

        m_date = re.search(r"Date:\s*(\d{1,2}\s+[A-Za-z]{3},\s*\d{4})", text)
        date_cert = "Unknown"
        if m_date:
            try:
                date_cert = datetime.strptime(m_date.group(1).replace(",", ""), "%d %b %Y").strftime("%d/%m/%Y")
            except Exception:
                pass

        # Parts (rules) - Using the robust regex pattern
        # Matches:
        #   CODE  TITLE  POINTS(or 'No')  STATUS  [ACHIEVED?]
        ROW_RE = re.compile(
            r"""^(?P<code>[A-Z][0-9]{2}\.\d)\s+
                (?P<title>.+?)\s+
                (?P<pts>(?:\d+(?:\.\d+)?(?:\+\d+(?:\.\d+)?)*)|No)\s+
                (?P<status>Achieved|Not\s+Attempted|Pending\s+Documentation|Not\s+Applicable)
                (?:\s+(?P<ach>(?:\d+(?:\.\d+)?(?:\+\d+(?:\.\d+)?)*)))?$""",
            re.VERBOSE | re.MULTILINE
        )

        def _sum_token(token: str):
            """Convert '1+1' or '0.5+1' or '2' to a float sum; return None if token invalid or 'No'."""
            if token is None:
                return None
            token = token.strip()
            if token.lower() == 'no':
                return None
            parts = token.split('+')
            try:
                return sum(float(p) for p in parts)
            except ValueError:
                return None

        def score_cell_from_line(line: str):
            m = ROW_RE.match(line.strip())
            if not m:
                return None  # let caller ignore or log
            
            pts_attempted = m.group('pts')
            status = m.group('status').replace('\xa0', ' ').strip()
            pts_ach = m.group('ach')

            attempted_val = _sum_token(pts_attempted)
            achieved_val = _sum_token(pts_ach)

            # Mapping according to the specified rules
            if status == 'Pending Documentation':
                return 'Pending Documentation'
            if status == 'Not Applicable':
                return 'Not Applicable'
            if status == 'Not Attempted':
                return None  # empty cell

            # Achieved
            if status == 'Achieved':
                # Prefer explicit achieved value if present
                if achieved_val is not None:
                    return achieved_val  # includes 0 or 0.5 etc.
                # Else fall back to attempted if numeric
                if attempted_val is not None:
                    return attempted_val
                # Truly no numeric score → 'p'
                return 'p'

            # Fallback (shouldn't happen)
            return None

        # Parse parts using the new robust scoring logic
        parts = []

        for m in ROW_RE.finditer(text):
            line = m.group(0)
            code = m.group('code')
            value = score_cell_from_line(line)
            if value is not None:
                parts.append({"code": code, "value": value})

        # Keep the old fallback, but also use finditer
        found_codes = {p["code"] for p in parts}

        PART_RE = re.compile(
            r"(?P<code>[AWNLVTSXMCI]\d{2}\.\d)\s+"
            r"(?P<title>.+?)\s+"
            r"(?:(?P<pre>\d+(?:\.\d+)?)\s+)?"
            r"(?P<status>Achieved|Not Attempted|Not Applicable|Withdrawn|"
            r"Pending(?: Documentation)?|Pending Documentation)"
            r"(?:\s+(?P<post>\d+(?:\.\d+)?))?",
            flags=re.IGNORECASE | re.DOTALL | re.MULTILINE
        )

        for m_old in PART_RE.finditer(text):
            code = m_old.group("code")
            if code in found_codes:
                continue
            status = (m_old.group("status") or "").strip().title()
            post = m_old.group("post")

            if status in ("Pending", "Pending Documentation"):
                value = "Pending Documentation"
            elif status == "Not Applicable":
                value = "Not Applicable"
            elif status == "Withdrawn":
                value = "Withdrawn"
            elif status == "Achieved":
                if post is not None:
                    try:
                        value = float(post)
                    except:
                        value = post
                else:
                    value = "p"
            else:  # Not Attempted
                value = None   # truly empty

            if value is not None:
                parts.append({"code": code, "value": value})

        # We deliberately DO NOT compute totals here; create_combined_excel will compute Sub-Points, Total, and %.

        return {
            "project_id": project_id,
            "project_name": project_name,
            "date_cert": date_cert,
            "parts": parts
        }

    except Exception as e:
        print(f"Error parsing WELL markdown: {e}")
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    try:
        if 'files' not in request.files:
            return jsonify({'error': 'No files provided'}), 400
        
        files = request.files.getlist('files')
        if not files or all(file.filename == '' for file in files):
            return jsonify({'error': 'No files selected'}), 400
        
        results = []
        
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file_path = os.path.join(UPLOAD_FOLDER, filename)
                file.save(file_path)
                
                # Store file size before processing
                file_size = os.path.getsize(file_path)
                
                # Extract first 2 pages
                first_two_pages_path, extract_error = extract_first_two_pages(file_path)
                if extract_error:
                    results.append({
                        'filename': filename,
                        'status': 'error',
                        'message': extract_error,
                        'file_size': file_size
                    })
                    continue
                
                # Convert to markdown
                markdown_content, convert_error = convert_to_markdown(first_two_pages_path)
                if convert_error:
                    results.append({
                        'filename': filename,
                        'status': 'error',
                        'message': convert_error,
                        'file_size': file_size
                    })
                    continue
                
                results.append({
                    'filename': filename,
                    'status': 'success',
                    'markdown': markdown_content,
                    'message': 'Successfully processed',
                    'file_size': file_size
                })
                
                # Clean up temporary files
                if os.path.exists(first_two_pages_path):
                    os.remove(first_two_pages_path)
                if os.path.exists(file_path):
                    os.remove(file_path)
        
        # Create combined Excel file
        excel_path, excel_error = create_combined_excel(results)
        if excel_error:
            return jsonify({'error': excel_error}), 500
        
        # Get just the filename for the client (avoid session storage)
        excel_filename = os.path.basename(excel_path)
        
        print(f"Debug: Excel file created at: {excel_path}")
        print(f"Debug: Excel filename: {excel_filename}")
        
        response_data = {
            'results': results,
            'excel_filename': excel_filename,
            'message': f'Successfully processed {len(results)} files. Combined Excel file created.'
        }
        
        print(f"Debug: Sending response to client: {response_data}")
        
        return jsonify(response_data)
        
    except Exception as e:
        # Log the error for debugging
        print(f"Upload error: {str(e)}")
        return jsonify({'error': f'Unexpected error during processing: {str(e)}'}), 500

@app.route('/download-excel')
def download_excel():
    """Download the combined Excel file using file parameter"""
    try:
        filename = request.args.get('file')
        if not filename:
            return jsonify({'error': 'Missing file parameter.'}), 400
        
        # Security: only allow files inside PROCESSED_FOLDER
        safe_name = os.path.basename(filename)
        excel_path = os.path.join(PROCESSED_FOLDER, safe_name)
        
        print(f"Debug: Downloading file: {safe_name} from path: {excel_path}")
        
        if not os.path.exists(excel_path):
            return jsonify({'error': f'Excel file not found: {safe_name}. Please process files first.'}), 404
        
        return send_file(
            excel_path, 
            as_attachment=True, 
            download_name=safe_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Download error: {str(e)}")
        return jsonify({'error': f'Error downloading file: {str(e)}'}), 500

@app.route('/clear-session', methods=['POST'])
def clear_session():
    """Clear session (kept for compatibility but simplified)"""
    try:
        session.clear()
        return jsonify({'message': 'Session cleared'})
    except Exception as e:
        return jsonify({'error': f'Error clearing session: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=DEBUG, host=HOST, port=PORT)
